import openpyxl
from openpyxl.utils import get_column_letter

# 새 워크북 생성
workbook = openpyxl.Workbook()

# 활성 시트 선택
sheet = workbook.active

# 헤더 추가
headers = ["ID", "이름", "수량", "가격"]
for col_num, header in enumerate(headers, 1):
    column_letter = get_column_letter(col_num)
    sheet[f"{column_letter}1"] = header

# 전자 제품 데이터 생성 (예시)
product_data = [
    (1, "노트북", 10, 800.0),
    (2, "스마트폰", 20, 400.0),
    # ... 여러 제품을 원하는 만큼 추가할 수 있습니다.
]

# 판매 목록에 데이터 추가
for row_idx, product in enumerate(product_data, start=2):
    for col_idx, value in enumerate(product, start=1):
        column_letter = get_column_letter(col_idx)
        sheet[f"{column_letter}{row_idx}"] = value

# 100개의 더미 행 추가 (예시)
for i in range(101, 201):
    sheet[f"A{i}"] = i  # ID
    sheet[f"B{i}"] = f"제품 {i}"  # 이름
    sheet[f"C{i}"] = 5  # 수량
    sheet[f"D{i}"] = 100.0  # 가격

# 파일로 저장
file_path = "c:\\work\\sales.xlsx"
workbook.save(file_path)

print(f"판매 목록이 {file_path} 경로에 저장되었습니다.")
