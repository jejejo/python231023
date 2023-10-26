import openpyxl

# 새 워크북 생성
workbook = openpyxl.Workbook()

# 활성 시트 선택
sheet = workbook.active

# 헤더를 열에 추가
sheet["A1"] = "ID"
sheet["B1"] = "이름"
sheet["C1"] = "수량"
sheet["D1"] = "가격"

# 전자 제품 데이터 생성
product_data = [
    (1, "세탁기", 10, 800.0),
    (2, "스마트폰", 20, 400.0),
    (3, "태블릿", 15, 300.0),
    (4, "헤드폰", 30, 50.0),
    (5, "스마트워치", 10, 150.0),
]

# 100개의 행으로 목록 채우기
for i in range(2, 102):
    product = product_data[i % 5]
    sheet.cell(row=i, column=1, value=product[0])  # ID
    sheet.cell(row=i, column=2, value=product[1])  # 이름
    sheet.cell(row=i, column=3, value=product[2])  # 수량
    sheet.cell(row=i, column=4, value=product[3])  # 가격

# 파일로 저장
file_path = "c:\\work\\sales2.xlsx"
workbook.save(file_path)

print(f"판매 목록이 {file_path}에 저장되었습니다.")
